import os
import pprint

import requests
import urllib3
import openpyxl
urllib3.disable_warnings()

def fetch_data():
    workbook=openpyxl.load_workbook("testdata.xlsx",data_only=True)
    sheet=workbook["NASA"]
    value = []
    for row in sheet.iter_rows(min_row=2):
        values=[]
        for cell in row:
            if cell.is_date:
                values.append(''.join(str(cell.value).split()[0]))
            else:
                values.append(cell.value)
        value.append(values)
    return value


def fetch_data_gemini():
    workbook=openpyxl.load_workbook("testdata.xlsx",data_only=True)
    sheet=workbook['Gemini']
    value=[]


    for row in range(2,sheet.max_row+1):
        column=1
        question=str(sheet.cell(row=row,column=column).value)
        expected=str(sheet.cell(row=row,column=column+1).value)
        value.append((question,expected))
    return value


def setup(start_date,end_date):
    api_key="pifPfdnfHkRmEK59TNGkKs8JRIPFry8TxU4h7mFv"
    url=f"https://api.nasa.gov/neo/rest/v1/feed?start_date={start_date}&end_date={end_date}"
    try:
        response=requests.get(url+f"&api_key={api_key}",verify=False).json()
        data=response.get("near_earth_objects").get(start_date)[2].get('estimated_diameter').get('feet')
        return data

    except Exception as e:
        raise RuntimeError("Exception occured due to ",e) from e


def gemini_api(API_KEY,question):
    import os
    import requests
    import json
    # API_KEY = os.getenv("GEMINI_API_KEY")
    if not API_KEY:
        raise RuntimeError("GEMINI_API_KEY not found")
    url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent"

    headers = {
        "Content-Type": "application/json"
    }
    payload = {
        "contents": [
            {
                "parts": [
                    {"text": question}
                ]
            }
        ]
    }
    params={"key":API_KEY}
    response=requests.post(url=url,headers=headers,params=params,json=payload,timeout=50)
    print(response.json())
    val = response.json().get('candidates')[0].get("content").get("parts")[0].get('text')
    print(val)
    return response

print(gemini_api(API_KEY=os.getenv("GEMINI_API_KEY"),question="what is 2+2"))